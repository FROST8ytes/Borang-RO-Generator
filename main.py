"""
This script automates the creation of Borang RO for quarantined patients of Malaysia.
It is licensed under GNU GPLv3. If possible, only modify the modifiable global variables.
I am not responsible of anything that happens to your machine if the script breaks something,
especially when you modify something that you shouldn't.
"""
# __author__ = "Ammar Farhan Mohamad Rizam"
# __copyright__ = "Copyright 2021, FROST8ytes"
# __date__ = "30.09.2021"
# __license__ = "GNU GPLv3"
# __version__ = "1.0"
# __emails__ = ["ammarfmr11@gmail.com", "amoh0096@student.monash.edu"]

from PyQt5.QtWidgets import QMainWindow, QApplication, QPushButton, QLabel, QFileDialog
from PyQt5 import uic
import sys

import openpyxl
import os
from datetime import datetime, timedelta
from docxtpl import DocxTemplate

###################################
#   MODIFIABLE GLOBAL VARIABLES   #
###################################
template_file_with_path = 'SAMPLE_WORD.docx'
excel_filename = 'SAMPLE_EXCEL.xlsx'
# sheets_to_read = ['SHEET1', 'SHEET2']
sheets_to_read = ['Sheet1']

quarantine_days_required = 10
doctors_name = "SAMPLE DOCTOR'S NAME"
doctors_position = "SAMPLE DOCTOR'S POSITION"
doctors_appointed_place = "SAMPLE APPOINTED PLACE"
appointed_place_phone_number = "012-3456789"
time = '9:30am'

###################################
# NON-MODIFIABLE GLOBAL VARIABLES #
###################################
names = []
identifications = []
addresses = []
phone_numbers = []
date_hsos = []
date_ros = []
source_sheet = []
todays_date = datetime.today()

###################################
#    END OF GLOBAL VARIABLES      #
###################################


###################################
#     DONT MODIFY CODE BELOW      #
#   UNLESS IF YOU KNOW WHAT YOU   #
#            ARE DOING            #
###################################

class UI(QMainWindow):
    def __init__(self):
        super(UI, self).__init__()

        # Load the ui file
        uic.loadUi("main_window.ui", self)

        self.file_name = ''
        
        # Define our widgets
        self.button_select_file = self.findChild(QPushButton, "btn_select_file")
        self.button_process = self.findChild(QPushButton, "btn_process")
        self.button_stop = self.findChild(QPushButton, "btn_stop")

        # Click the btn to select file
        self.button_select_file.clicked.connect(self.click_select_file)

        # Click the btn to process
        self.button_process.clicked.connect(self.click_process)

        # Click the btn to stop
        self.button_stop.clicked.connect(self.click_stop)

        # Show The App
        self.show()

    # parse_data function
    def parse_data(self, excel_file: openpyxl.Workbook, sheet_name: str):
        print('parse data')
        
        """
        This function opens the sheet_name from the opened excel file, and parses
        - patient's name
        - patient's identification (MyKad/Passport)
        - patient's address (maybe employer's address)
        - patient's phone number (maybe employer's phone number)
        - patient's first date of quarantine
        - patient's last date of quarantine

        :param excel_file: xlsx file to read from
        :param sheet_name: sheet name from xlsx file to read from
        """
        global names, identifications, addresses, phone_numbers, date_hsos, date_ros, source_sheet
        
        sheet = excel_file[sheet_name]

        print(f'[*] Parsing data from sheet {sheet_name}...')
        
        for row in range(2, sheet.max_row + 1):
            source_sheet.append(sheets_to_read.index(sheet_name))

            name = sheet['B' + str(row)].value
            if name is None:
                break
            names.append(name)
            print(f'Adding name: {name} from sheet {sheet_name} to record')

            identification_number = sheet['C' + str(row)].value
            identifications.append(identification_number)
            print(f"Adding identification number: {identification_number} from sheet {sheet_name} to record")

            address = sheet['F' + str(row)].value
            addresses.append(address)
            print(f'Adding work address: {address} from sheet {sheet_name} to record')

            phone_number = sheet['E' + str(row)].value
            phone_numbers.append(phone_number)
            print(f'Adding phone number: {phone_number} from sheet {sheet_name} to record')

            date_hso = sheet['G' + str(row)].value
            date_hsos.append(date_hso)
            print(f'Adding date HSO: {date_hso} from sheet {sheet_name} to record')

            date_ro = sheet['H' + str(row)].value
            date_ros.append(date_ro)
            print(f'Adding date RO: {date_ro} from sheet {sheet_name} to record')

        print(f'[+] Data from sheet {sheet_name} parsed successfully!')


    # generate_docx function
    def generate_docx(self, generated_name_with_path: str, name: str, identification_number: str, address: str, phone_number: str, date_hso: str, date_ro: str):
        """
        This function generates a docx file based on the template of Borang RO. Refer to SAMPLE_WORD.docx for an example
        of the template.

        :param generated_name_with_path: name of the file that is generated
        :param name: name of patient
        :param identification_number: MyKad/Passport number of patient
        :param address: home/work address of patient (most likely employer's address)
        :param phone_number: phone number of patient/employer
        :param date_hso: date of first day of quarantine dd/mm/yyyy
        :param date_ro: date of last day of quarantine dd/mm/yyyy
        """
        print('generate docx')
        
        # global template_file_with_path, todays_date, time, doctors_name, doctors_position, \
        #     doctors_appointed_place, appointed_place_phone_number
        
        # doc = DocxTemplate(template_file_with_path if template_file_with_path.endswith('.docx')
        #                     else f"{template_file_with_path}.docx")

        # context = {'name': name.upper(), 'identification_number': identification_number,
        #             'address': address.upper() if address is not None else 'NO GIVEN ADDRESS',
        #             'phone_number': phone_number if phone_number is not None else 'N/A',
        #             'date_hso': date_hso if date_hso is not None else 'NO GIVEN DATE',
        #             'date_ro': date_ro if date_ro is not None else 'NO GIVEN DATE',
        #             'doctors_name': doctors_name.upper(),
        #             'doctors_position': doctors_position.upper(),
        #             'appointed_place': doctors_appointed_place.upper(),
        #             'appointed_place_phone': appointed_place_phone_number,
        #             'todays_date': todays_date.strftime('%d/%m/%Y'),
        #             'time': time.upper()
        #             }

        # doc.render(context)
        # print(f'[*] Generating file {generated_name_with_path}...')
        # doc.save(generated_name_with_path if generated_name_with_path.endswith('.docx')
        #         else f"{generated_name_with_path}.docx")
        # print(f'[+] File {generated_name_with_path} generated successfully!')
    
    
    # btn to select file
    def click_select_file(self):
        # self.label_select_file.setText("btn 1")
        # Open File Dialog
        fname = QFileDialog.getOpenFileName(self, "Open File", "", "All Files (*);;XLS Files (*.xls);; XLSX Files (*.xlsx)" )
        
        self.file_name = fname[0]
        
        print(fname[0])
    
    
    # btn to process main task
    def click_process(self):
        print('click Process')
        print(f'file name: {self.file_name}')
        global names, \
            identifications, \
            addresses, \
            phone_numbers, \
            date_hsos, \
            date_ros, \
            source_sheet, \
            sheets_to_read, \
            todays_date, \
            time \
            # excel_filename
        
        # excel_file = openpyxl.load_workbook(excel_filename if excel_filename.endswith('.xlsx') else excel_filename + '.xlsx')
        excel_file = openpyxl.load_workbook(self.file_name)
        
        # sheet = excel_file['Sheet1']

        # print(f'[*] Parsing data from sheet {sheet}...')
        
        for sheet in sheets_to_read:
            # self.parse_data(excel_file, sheet)
            
            print(f'[*] Creating folder {sheet}...')
            try:
                os.mkdir(sheet)
                print(f'[+] Folder {sheet} created successfully!')
            except FileExistsError:
                print(f'[-] Folder {sheet} exists.')

        # for i in range(len(names)):
        #     days_since_first_quarantined = todays_date - datetime.strptime(date_hsos[i], "%d/%m/%Y")
        #     if days_since_first_quarantined >= timedelta(days=quarantine_days_required):
        #         self.generate_docx(f"{os.getcwd()}/{sheets_to_read[source_sheet[i]]}/"
        #                     f"{i:02}-Borang_RO-{names[i].replace('/', '')}.docx",
        #                     names[i], identifications[i], addresses[i], phone_numbers[i], date_hsos[i], date_ros[i])

    def click_stop(self):
    # self.label_stop.setText("btn 3")
        print('click stop')

# # Initialize The App
app = QApplication(sys.argv)
UIWindow = UI()
app.exec_()
