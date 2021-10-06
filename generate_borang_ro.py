#!/usr/bin/env python3

"""
This script automates the creation of Borang RO for quarantined patients of Malaysia.
It is licensed under GNU GPLv3. If possible, only modify the modifiable global variables.
I am not responsible of anything that happens to your machine if the script breaks something,
especially when you modify something that you shouldn't.
"""
__author__ = "Ammar Farhan Mohamad Rizam"
__copyright__ = "Copyright 2021, FROST8ytes"
__date__ = "30.09.2021"
__license__ = "GNU GPLv3"
__version__ = "1.0"
__emails__ = ["ammarfmr11@gmail.com", "amoh0096@student.monash.edu"]

import openpyxl
import os
import argparse
from sys import platform
from datetime import datetime, timedelta
from docxtpl import DocxTemplate

###################################
#   MODIFIABLE GLOBAL VARIABLES   #
###################################
template_file_with_path = 'SAMPLE_WORD.docx'
excel_filename = 'SAMPLE_EXCEL.xlsx'
sheets_to_read = ['SHEET1', 'SHEET2']
quarantine_days_required = 10

# TODO: Implement sqlite which contains the next following info below
doctors_name = "SAMPLE DOCTOR'S NAME"
doctors_position = "SAMPLE DOCTOR'S POSITION"
doctors_appointed_place = "SAMPLE APPOINTED PLACE"
appointed_place_phone_number = "012-3456789"

###################################
#    END OF GLOBAL VARIABLES      #
###################################


###################################
#     DONT MODIFY CODE BELOW      #
#   UNLESS IF YOU KNOW WHAT YOU   #
#            ARE DOING            #
###################################
class ExcelParserAndDocxGenerator:

    # TODO: Add docstring for each method

    def __init__(self, excel_file_name_with_path: str, sheet_names: list[str],
                 template_file_name_with_path: str = None) -> None:
        self.excel_file: openpyxl.Workbook = openpyxl.load_workbook(excel_file_name_with_path
                                                                    if excel_filename.endswith('.xlsx')
                                                                    else excel_filename + '.xlsx'
                                                                    )
        self.sheet_names: list[str] = sheet_names
        self.patients: dict = {}
        self.base_directory = os.path.abspath(os.path.dirname(__file__))
        self.sheet = None
        self.template_file: DocxTemplate = self.open_and_set_template_file(template_file_name_with_path)

    def open_and_set_template_file(self, template_file_name_with_path: str = None) -> DocxTemplate:
        if template_file_name_with_path is not None:
            return DocxTemplate(template_file_name_with_path if template_file_name_with_path.endswith('.docx')
                                else f"{template_file_name_with_path}.docx")
        else:
            return DocxTemplate(self.base_directory + '/SAMPLE_WORD.docx')

    def set_template_file(self, template_file_name_with_path: str = None):
        self.template_file = self.open_and_set_template_file(template_file_name_with_path)

    def get_available_sheet_names_from_excel_file(self) -> list[str]:
        return self.excel_file.sheetnames

    def make_directory(self, sheet_name: str) -> str:
        print(f"[*] Creating folder {sheet_name}...")
        try:
            os.mkdir(f"{self.base_directory}/{sheet_name}")
            print(f"[+] Folder {sheet_name} created successfully!")
        except FileExistsError:
            print(f"[-] Folder {sheet_name} exists.")
            print(f"[*] Using existing folder...")

        return f"{self.base_directory}\\{sheet_name}" if platform.startswith('win32') \
            else f"{self.base_directory}/{sheet_name}"

    def read_parse_and_generate_all(self) -> None:
        self.parse_all()
        # TODO: Add parsed data to sqlite to keep track of who's letter needs to be generated
        self.generate_all_docx()

    def parse_all(self):
        for sheet_name in self.sheet_names:
            self.patients[sheet_name] = self.parse_patients_in_sheet(sheet_name)

    def parse_patients_in_sheet(self, sheet_name: str) -> list[dict]:
        self.sheet = self.excel_file[sheet_name]
        patients: list[dict] = []

        print(f"[*] Parsing data from sheet {sheet_name}...")
        # TODO: detect which row contains column headers
        for row in range(2, self.sheet.max_row + 1):
            patients.append(self.parse_patient_info(row))
            print(f"[+] Data from sheet {sheet_name} parsed successfully!")

        return patients

    def parse_patient_info(self, row: int) -> dict:
        patient_info = {}
        name = self.sheet['B' + str(row)].value
        if name is None:
            raise TypeError("[-] Patient has no name.")

        # TODO: Make columns more flexible
        patient_info["name"] = name
        print(f"Adding name: {name} to record")
        identification_number = self.sheet['C' + str(row)].value
        patient_info["id"] = identification_number
        print(f"Adding identification number: {identification_number} to record")
        address = self.sheet['F' + str(row)].value
        patient_info["address"] = address
        print(f"Adding work address: {address} to record")
        phone_number = self.sheet['E' + str(row)].value
        patient_info["phone_number"] = phone_number
        print(f"Adding phone number: {phone_number} to record")
        date_hso = self.sheet['G' + str(row)].value
        patient_info["date_hso"] = date_hso
        print(f"Adding date HSO: {date_hso} to record")
        date_ro = self.sheet['H' + str(row)].value
        patient_info["date_ro"] = date_ro
        print(f"Adding date RO: {date_ro} to record")

        return patient_info

    def generate_all_docx(self, signed_date_time: datetime = datetime.today()) -> None:
        # TODO: figure out efficient ways to replace nested for-loops
        for sheet_name in self.patients.keys():
            new_directory = self.make_directory(sheet_name)
            print(f"[*] Generating {new_directory} forms...")
            for patients in self.patients[sheet_name]:
                for patient in patients:
                    self.generate_docx(new_directory, patient, signed_date_time)
            print(f"[+] Successfully generated {new_directory} forms!")

    def generate_docx(self, path: str, patient_info: dict, signed_date_time: datetime = datetime.today()) -> None:
        global doctors_name, doctors_position, doctors_appointed_place,\
            appointed_place_phone_number, quarantine_days_required

        days_since_first_quarantined = signed_date_time - datetime.strptime(patient_info["date_hso"], "%d/%m/%Y")
        if days_since_first_quarantined >= timedelta(days=quarantine_days_required):
            print(f"[-] {patient_info['name']} still under quarantine...")
            print(f"[*] Form for {patient_info['name']} not generated.")
            return

        context = {'name': patient_info["name"].upper(),
                   'identification_number': patient_info["identification_number"],
                   'address': patient_info["address"].upper() if patient_info["address"] is not None else "NO GIVEN ADDRESS",
                   'phone_number': patient_info["phone_number"] if patient_info["phone_number"] is not None else "N/A",
                   'date_hso': patient_info["date_hso"] if patient_info["date_hso"] is not None else "NO GIVEN DATE",
                   'date_ro': patient_info["date_ro"] if patient_info["date_ro"] is not None else "NO GIVEN DATE",
                   'doctors_name': doctors_name.upper(),
                   'doctors_position': doctors_position.upper(),
                   'appointed_place': doctors_appointed_place.upper(),
                   'appointed_place_phone': appointed_place_phone_number,
                   'todays_date': signed_date_time.strftime('%d/%m/%Y'),
                   'time': signed_date_time.strftime("%I:%M%p")
                   }

        self.template_file.render(context)
        generated_file_name = path if path.endswith('.docx') else f"{path}/{patient_info['name'].replace('/', '')}.docx"
        print(f"[*] Generating file {generated_file_name}...")
        self.template_file.save(generated_file_name)
        print(f"[+] File {generated_file_name} generated successfully!")


def main():
    global excel_filename, template_file_with_path, sheets_to_read
    parser = argparse.ArgumentParser(description="read excel spreadsheet and/or generate RO form", add_help=True)
    parser.add_argument("-v", "--verbose", help="print out messages (helpful when debugging)")
    parser.add_argument("-s", "--source-file", help="specify source file to parse (.xlsx format)")
    parser.add_argument("-t", "--template-file", help="specify template file to generate from (.docx format)")
    # TODO: complete options and implement them into app instead of using default global variables' values
    args = parser.parse_args()
    app = ExcelParserAndDocxGenerator(excel_filename, sheets_to_read, template_file_with_path)
    app.generate_all_docx()


if __name__ == '__main__':
    main()
